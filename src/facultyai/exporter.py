"""Excel exporter: writes ``faculty_data.xlsx`` from the database.

Supports incremental updates: when called with a specific university + department,
removes old rows for that combo and appends new results. Otherwise writes all
active records fresh.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from .database import Database
from .schema import Schema


def _col_letter(index: int) -> str:
    return get_column_letter(index + 1)


def _resolve_formula(formula: str, column_index: dict[str, int]) -> str:
    def _replace(m: re.Match[str]) -> str:
        col_name = m.group(1)
        idx = column_index.get(col_name)
        if idx is None:
            return f"[@{col_name}]"
        return _col_letter(idx) + "{row}"

    return re.sub(r"\[@\[([^\]]+)\]\]", _replace, formula)


async def export_to_excel(
    db: Database,
    schema: Schema,
    output_path: str | Path,
    *,
    upsert_university: str | None = None,
    upsert_department: str | None = None,
) -> int:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    headers = [c.name for c in schema.columns]
    col_map = {c.name: i for i, c in enumerate(schema.columns)}
    formula_cols = [(col_map[c.name], c) for c in schema.columns if c.is_formula()]
    static_cols = [(col_map[c.name], c) for c in schema.columns if c.is_static()]

    if upsert_university is not None:
        db_rows = await db.get_faculty_by_university(upsert_university, upsert_department)
    else:
        db_rows = await db.get_active_faculty()

    new_rows = _rows_from_db(db_rows, headers, col_map, schema, static_cols, formula_cols)

    if upsert_university is not None and path.exists():
        dedup_values: dict[str, str] = {}
        for key in schema.dedup_keys:
            if key == "Institution":
                dedup_values[key] = upsert_university
            elif key == "Department" and upsert_department:
                dedup_values[key] = upsert_department
            # For other dedup keys, values come from the new rows
        _merge_incremental(path, headers, new_rows, dedup_values)
        return len(new_rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in new_rows:
        ws.append(row)
    wb.save(path)

    return len(new_rows)


def _rows_from_db(
    db_rows: list[dict[str, Any]],
    headers: list[str],
    col_map: dict[str, int],
    schema: Schema,
    static_cols: list[tuple[int, Any]],
    formula_cols: list[tuple[int, Any]],
) -> list[list[Any]]:
    new_data: list[list[Any]] = []
    for row_offset, r in enumerate(db_rows):
        parsed = json.loads(r["data_json"] or "{}")
        parsed["university_name"] = r["university"]
        parsed["department"] = r["department"]

        # Excel row number: header is row 1, so first data row is row 2
        excel_row = row_offset + 2

        record: list[Any] = [None] * len(headers)
        for col in schema.extracted_columns():
            idx = col_map[col.name]
            val = parsed.get(col.name, "")
            record[idx] = val if val else ""
        for idx, col in static_cols:
            if col.value_from:
                record[idx] = parsed.get(col.value_from, "")
            else:
                record[idx] = col.value or ""
        for idx, col in formula_cols:
            raw = _resolve_formula(col.formula or "", col_map)
            record[idx] = raw.replace("{row}", str(excel_row))
        new_data.append(record)
    return new_data


def _merge_incremental(
    path: Path,
    headers: list[str],
    new_rows: list[list[Any]],
    dedup_values: dict[str, str],
) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]] if wb.sheetnames else wb.active

    # Find column indices for dedup keys in the existing file
    key_cols: dict[str, int | None] = {}
    for key in dedup_values:
        key_cols[key] = _find_col(ws, key)

    existing: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_list = list(row)
        match = True
        for key, expected in dedup_values.items():
            col = key_cols.get(key)
            cell_val = ""
            if col is not None and col < len(row_list) and row_list[col] is not None:
                cell_val = str(row_list[col]).strip()
            if cell_val != expected.strip():
                match = False
                break
        if match:
            continue  # Remove stale row
        existing.append(row_list)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(headers)
    for row in existing:
        ws2.append(row)
    for row in new_rows:
        ws2.append(row)
    wb2.save(path)
    wb.close()


def _find_col(ws, name: str) -> int | None:
    """Return 0-based column index for a header name, or None if not found."""
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val and str(val).strip().lower() == name.lower():
            return c - 1
    return None
