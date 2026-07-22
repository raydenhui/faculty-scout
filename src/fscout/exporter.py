"""Excel exporter: writes ``faculty_data.xlsx`` from the database.

Supports incremental updates: when called with a specific university + department,
removes old rows for that combo (matched by _source_key) and appends new results.
Otherwise writes all active records fresh.

The ``_source_key`` column is a hidden internal column containing the
``{department}/{university}`` values from the input file (not LLM-extracted).
It enables reliable deduplication regardless of schema configuration.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from .database import Database
from .schema import Schema

_SOURCE_KEY_HEADER = "_source_key"


def _col_letter(index: int) -> str:
    return get_column_letter(index + 1)


def _resolve_formula(formula: str, column_index: dict[str, int]) -> str:
    def _replace(m: re.Match[str]) -> str:
        col_name = m.group(1)
        idx = column_index.get(col_name)
        if idx is None:
            return f"[@{col_name}]"
        return _col_letter(idx) + "{row}"

    return re.sub(r"\[@\[([^\]]+)\]\]", _replace, formula)


async def export_to_excel(
    db: Database,
    schema: Schema,
    output_path: str | Path,
    *,
    upsert_university: str | None = None,
    upsert_department: str | None = None,
) -> int:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    headers = [_SOURCE_KEY_HEADER] + [c.name for c in schema.columns]
    col_map = {c.name: i + 1 for i, c in enumerate(schema.columns)}
    col_map[_SOURCE_KEY_HEADER] = 0
    formula_cols = [(col_map[c.name], c) for c in schema.columns if c.is_formula()]
    static_cols = [(col_map[c.name], c) for c in schema.columns if c.is_static()]

    if upsert_university is not None:
        db_rows = await db.get_faculty_by_university(upsert_university, upsert_department)
    else:
        db_rows = await db.get_active_faculty()

    source_key = _build_source_key(upsert_university, upsert_department)
    new_rows = _rows_from_db(db_rows, headers, col_map, schema, static_cols, formula_cols, source_key)

    if upsert_university is not None and path.exists():
        _merge_incremental(path, headers, new_rows, source_key)
        return len(new_rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in new_rows:
        ws.append(row)
    _hide_source_key_column(ws)
    wb.save(path)

    return len(new_rows)


def _build_source_key(university: str | None, department: str | None) -> str:
    return f"{department or ''}/{university or ''}"


def _rows_from_db(
    db_rows: list[dict[str, Any]],
    headers: list[str],
    col_map: dict[str, int],
    schema: Schema,
    static_cols: list[tuple[int, Any]],
    formula_cols: list[tuple[int, Any]],
    source_key: str = "",
) -> list[list[Any]]:
    new_data: list[list[Any]] = []
    for row_offset, r in enumerate(db_rows):
        parsed = json.loads(r["data_json"] or "{}")
        parsed["university_name"] = r["university"]
        parsed["department"] = r["department"]

        excel_row = row_offset + 2

        record: list[Any] = [None] * len(headers)
        record[0] = _build_source_key(r["university"], r.get("department"))
        for col in schema.extracted_columns():
            idx = col_map[col.name]
            val = parsed.get(col.name, "")
            record[idx] = val if val else ""
        for col in schema.fallback_columns():
            idx = col_map[col.name]
            val = parsed.get(col.name, "")
            record[idx] = val if val else ""
        for idx, col in static_cols:
            if col.value_from:
                record[idx] = parsed.get(col.value_from, "")
            else:
                record[idx] = col.value or ""
        for idx, col in formula_cols:
            raw = _resolve_formula(col.formula or "", col_map)
            record[idx] = raw.replace("{row}", str(excel_row))
        new_data.append(record)
    return new_data


def _merge_incremental(
    path: Path,
    headers: list[str],
    new_rows: list[list[Any]],
    source_key: str,
) -> None:
    """Remove rows matching *source_key*, keep the rest, append *new_rows*."""
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]] if wb.sheetnames else wb.active

    sk_col = _find_col(ws, _SOURCE_KEY_HEADER)

    existing: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_list = list(row)
        if _source_key_matches(row_list, sk_col, source_key):
            continue
        existing.append(row_list)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(headers)
    for row in existing:
        ws2.append(row)
    for row in new_rows:
        ws2.append(row)
    _hide_source_key_column(ws2)
    wb2.save(path)
    wb.close()


def _source_key_matches(row: list[Any], sk_col: int | None, expected: str) -> bool:
    if sk_col is None:
        return False
    if sk_col >= len(row):
        return False
    cell_val = row[sk_col]
    if cell_val is None:
        return False
    return str(cell_val).strip() == expected.strip()


def _hide_source_key_column(ws) -> None:
    """Hide the _source_key column (column A) via openpyxl after writing."""
    if ws.max_column >= 1:
        letter = _col_letter(0)
        ws.column_dimensions[letter].hidden = True


def _find_col(ws, name: str) -> int | None:
    """Return 0-based column index for a header name, or None if not found."""
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val and str(val).strip().lower() == name.lower():
            return c - 1
    return None


def export_records(
    records: list[dict[str, Any]],
    schema: Schema,
    output_path: Path,
    *,
    source_university: str = "",
    source_department: str = "",
) -> None:
    """Write *records* directly to Excel, merging incrementally by _source_key.

    This is the DB-free counterpart to ``export_to_excel``.  It accepts
    in-memory records and a source university/department key so it can
    replace stale rows for the same department and append fresh data.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    headers = [_SOURCE_KEY_HEADER] + [c.name for c in schema.columns]
    col_map = {c.name: i + 1 for i, c in enumerate(schema.columns)}
    col_map[_SOURCE_KEY_HEADER] = 0
    formula_cols = [(col_map[c.name], c) for c in schema.columns if c.is_formula()]
    static_cols = [(col_map[c.name], c) for c in schema.columns if c.is_static()]

    source_key = _build_source_key(source_university, source_department)
    new_rows = _rows_from_records(records, headers, col_map, schema, static_cols, formula_cols,
                                  source_university, source_department)

    if output_path.exists() and source_key:
        _merge_incremental(output_path, headers, new_rows, source_key)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in new_rows:
            ws.append(row)
        _hide_source_key_column(ws)
        wb.save(output_path)


def _rows_from_records(
    records: list[dict[str, Any]],
    headers: list[str],
    col_map: dict[str, int],
    schema: Schema,
    static_cols: list[tuple[int, Any]],
    formula_cols: list[tuple[int, Any]],
    university: str,
    department: str | None,
) -> list[list[Any]]:
    """Build Excel rows from in-memory records (no DB lookup)."""
    new_data: list[list[Any]] = []
    for row_offset, rec in enumerate(records):
        excel_row = row_offset + 2

        row: list[Any] = [None] * len(headers)
        row[0] = _build_source_key(university, department)

        for col in schema.extracted_columns():
            idx = col_map[col.name]
            val = rec.get(col.name, "")
            row[idx] = val if val else ""
        for col in schema.fallback_columns():
            idx = col_map[col.name]
            val = rec.get(col.name, "")
            row[idx] = val if val else ""
        for idx, col in static_cols:
            if col.value_from:
                meta = {"university_name": university, "department": department or ""}
                row[idx] = meta.get(col.value_from, "")
            else:
                row[idx] = col.value or ""
        for idx, col in formula_cols:
            raw = _resolve_formula(col.formula or "", col_map)
            row[idx] = raw.replace("{row}", str(excel_row))
        new_data.append(row)
    return new_data
